import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from tkcalendar import DateEntry
import cv2
import pyzbar.pyzbar as pyzbar
import openpyxl as xl
import numpy as np
from datetime import datetime

def select_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        from_date = datetime.combine(from_date_entry.get_date(), datetime.min.time())
        to_date = datetime.combine(to_date_entry.get_date(), datetime.max.time())
        process_images(folder_path, from_date, to_date)
        
def process_images(folder_path, from_date, to_date):
    image_extensions = [".tif", ".tiff"]
    files = [f for f in os.listdir(folder_path) if any(f.lower().endswith(ext) for ext in image_extensions)]

    # Filtrar archivos por fecha de modificación
    filtered_files = []
    for filename in files:
        file_path = os.path.join(folder_path, filename)
        file_mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))  # Corregido
        if from_date <= file_mod_time <= to_date:
            filtered_files.append(filename)

    total_files = len(filtered_files)
    progress_bar['maximum'] = total_files

    wb, ws = create_excel_workbook("Codigos de Barra", ["Receta", "Autorizacion", "Ruta Archivo"])
    errores_wb, errores_ws = create_excel_workbook("Errores", ["Archivo", "Ruta Completa"])

    recetas = []

    for i, filename in enumerate(filtered_files):
        image_path = os.path.join(folder_path, filename)
        barcodes = read_barcodes(image_path)
        if barcodes:
            ws.append(barcodes + [image_path])
            recetas.append(barcodes[0])
        else:
            errores_ws.append([filename, image_path])
        update_progress_bar(i + 1)

    open_error_files(errores_ws, folder_path)
    handle_manual_entries(errores_ws, ws, recetas)
    save_workbooks(wb, errores_wb, recetas)
    
    progress_bar['value'] = 0

def create_excel_workbook(sheet_title, headers):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    return wb, ws

def read_barcodes(image_path):
    try:
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        height, width = image.shape

        preprocessings = [
            lambda img: img,
            lambda img: cv2.resize(img, (width * 2, height * 2)),
            lambda img: cv2.GaussianBlur(img, (5, 5), 0),
            lambda img: cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2),
            lambda img: cv2.threshold(img, 128, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1],
            lambda img: cv2.Canny(img, 100, 200)
        ]

        for preprocess in preprocessings:
            decoded_codes = decode_barcodes(preprocess(image))
            if decoded_codes:
                return decoded_codes

        additional_preprocessings = [
            lambda img: cv2.equalizeHist(img),
            lambda img: cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 11, 2),
            lambda img: cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1],
            lambda img: cv2.morphologyEx(img, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5)))
        ]

        for preprocess in additional_preprocessings:
            decoded_codes = decode_barcodes(preprocess(image))
            if decoded_codes:
                return decoded_codes

        return handle_image_rotation(image, preprocessings + additional_preprocessings)

    except Exception as e:
        print(f"Error al procesar la imagen {image_path}: {e}")
        return []

def decode_barcodes(processed_image):
    barcodes = pyzbar.decode(processed_image)
    decoded_codes = []

    for barcode in sorted(barcodes, key=lambda x: x.rect.top):
        for encoding in ["utf-8", "latin1", "ascii"]:
            try:
                barcode_data = barcode.data.decode(encoding)
                if len(barcode_data) == 13:
                    decoded_codes.append(barcode_data)
                    break
            except UnicodeDecodeError:
                continue

    return decoded_codes if len(decoded_codes) >= 2 else None

def handle_image_rotation(image, preprocessings):
    height, width = image.shape
    edges = cv2.Canny(image, 100, 200)
    lines = cv2.HoughLinesP(edges, 1, np.pi/180, 100, minLineLength=100, maxLineGap=10)

    if lines is not None:
        angles = [np.arctan2(y2 - y1, x2 - x1) * 180 / np.pi for x1, y1, x2, y2 in lines[:, 0]]
        angle = np.median(angles)
        M = cv2.getRotationMatrix2D((width // 2, height // 2), angle, 1)
        rotated = cv2.warpAffine(image, M, (width, height))
        top_crop = int(height * 0.25)
        rotated = rotated[top_crop:, :]

        for preprocess in preprocessings:
            decoded_codes = decode_barcodes(preprocess(rotated))
            if decoded_codes:
                return decoded_codes

    return []

def update_progress_bar(value):
    progress_bar['value'] = value
    root.update_idletasks()

def handle_manual_entries(errores_ws, ws, recetas):
    if messagebox.askyesno("Proceso completado", "Desea cargar manualmente los valores de recetas para los errores antes de continuar?"):
        for row in errores_ws.iter_rows(min_row=2, values_only=True):
            filename, image_path = row
            receta = simpledialog.askstring("Entrada manual", f"Ingrese el código de receta para {filename}:")
            autorizacion = simpledialog.askstring("Entrada manual", f"Ingrese el código de autorización para {filename}:")
            if receta and autorizacion:
                ws.append([receta, autorizacion, image_path])
                recetas.append(receta)

def save_workbooks(wb, errores_wb, recetas):
    excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if excel_path:
        wb.save(excel_path)
        messagebox.showinfo("Proceso completado", f"Se ha generado el archivo Excel en:\n{excel_path}")

    txt_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    if txt_path:
        save_recetas_to_txt(txt_path, recetas)
        messagebox.showinfo("Proceso completado", f"Se ha generado el archivo de texto en:\n{txt_path}")

    errores_excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if errores_excel_path:
        errores_wb.save(errores_excel_path)
        messagebox.showinfo("Proceso completado", f"Se ha generado el archivo Excel de errores en:\n{errores_excel_path}")

def save_recetas_to_txt(txt_path, recetas):
    try:
        with open(txt_path, 'w') as f:
            for receta in recetas:
                f.write(receta + "\n")
    except Exception as e:
        print(f"Error al guardar el archivo de texto {txt_path}: {e}")

def open_error_files(errores_ws, folder_path):
    errores_tif_files = [filename for filename, _ in errores_ws.iter_rows(min_row=2, values_only=True)
                        if filename.lower().endswith((".tif", ".tiff"))]
    if errores_tif_files and messagebox.askyesno("Abrir archivos de errores", "¿Desea abrir los archivos .tif de errores ahora?"):
        for filename in errores_tif_files:
            open_tif_file(os.path.join(folder_path, filename))

def open_tif_file(image_path):
    try:
        if os.path.exists(image_path):
            os.startfile(image_path)
    except Exception as e:
        print(f"No se pudo abrir la imagen {image_path}: {e}")

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Extractor de Códigos de Barra")
root.geometry("600x500")
root.config(background="#2c3e50")

style = ttk.Style()
style.configure("TButton", font=("Helvetica", 12, "bold"), padding=10, background="#3498db", foreground="black")
style.map("TButton", background=[('active', '#2980b9')])

select_folder_button = ttk.Button(root, text="Seleccionar Carpeta", command=select_folder)
select_folder_button.pack(pady=20)

from_label = ttk.Label(root, text="Desde:", font=("Helvetica", 12, "bold"), background="#2c3e50", foreground="white")
from_label.pack(pady=5)
from_date_entry = DateEntry(root, font=("Helvetica", 12), background="blue", foreground="white", borderwidth=2)
from_date_entry.pack(pady=5)

to_label = ttk.Label(root, text="Hasta:", font=("Helvetica", 12, "bold"), background="#2c3e50", foreground="white")
to_label.pack(pady=5)
to_date_entry = DateEntry(root, font=("Helvetica", 12), background="blue", foreground="white", borderwidth=2)
to_date_entry.pack(pady=5)

progress_bar = ttk.Progressbar(root, orient=tk.HORIZONTAL, length=400, mode='determinate')
progress_bar.pack(pady=20)

root.mainloop()
